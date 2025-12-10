import customtkinter as ctk
from tkinter import filedialog, messagebox
import pandas as pd
import openpyxl
import os
import datetime

class CustomMessageBox:
    def __init__(self, parent, title, message):
        self.dialog = ctk.CTkToplevel(parent)
        self.dialog.title(title)

        frame = ctk.CTkFrame(self.dialog)
        frame.pack(expand=True, fill="both", padx=20, pady=20)

        label = ctk.CTkLabel(frame, text=message, wraplength=400)
        label.pack(pady=10)

        button_frame = ctk.CTkFrame(frame)
        button_frame.pack(pady=10)

        ok_button = ctk.CTkButton(button_frame, text="OK", command=self.dialog.destroy)
        ok_button.pack()

        self.dialog.transient(parent)
        self.dialog.grab_set()
        parent.wait_window(self.dialog)


class InvoiceToolApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("自动填写开票模板工具 (合并版)")
        self.geometry("800x450")
        
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")

        # Main frame
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)

        # File selection frame
        self.file_frame = ctk.CTkFrame(self)
        self.file_frame.grid(row=0, column=0, padx=20, pady=10, sticky="ew")
        self.file_frame.grid_columnconfigure(1, weight=1)

        self.source_path = ctk.StringVar()
        self.template_path = ctk.StringVar()

        ctk.CTkLabel(self.file_frame, text="1. 选择源数据文件 (上研-满座儿.xlsx):").grid(row=0, column=0, columnspan=3, padx=10, pady=5, sticky="w")
        ctk.CTkEntry(self.file_frame, textvariable=self.source_path).grid(row=1, column=0, columnspan=2, padx=10, pady=5, sticky="ew")
        ctk.CTkButton(self.file_frame, text="浏览", command=self.select_source).grid(row=1, column=2, padx=10, pady=5)

        ctk.CTkLabel(self.file_frame, text="2. 选择开票模板文件 (导入开票模板.xlsx):").grid(row=2, column=0, columnspan=3, padx=10, pady=5, sticky="w")
        ctk.CTkEntry(self.file_frame, textvariable=self.template_path).grid(row=3, column=0, columnspan=2, padx=10, pady=5, sticky="ew")
        ctk.CTkButton(self.file_frame, text="浏览", command=self.select_template).grid(row=3, column=2, padx=10, pady=5)

        # Parameters frame
        self.params_frame = ctk.CTkFrame(self)
        self.params_frame.grid(row=1, column=0, padx=20, pady=10, sticky="ew")
        self.params_frame.grid_columnconfigure((1, 3, 5), weight=1)

        ctk.CTkLabel(self.params_frame, text="税收编码 (餐饮):").grid(row=0, column=0, padx=10, pady=10, sticky="w")
        self.tax_code_entry = ctk.CTkEntry(self.params_frame)
        self.tax_code_entry.insert(0, "3070401000000000000") # 默认餐饮编码
        self.tax_code_entry.grid(row=0, column=1, padx=10, pady=10, sticky="ew")

        ctk.CTkLabel(self.params_frame, text="税率 (小数):").grid(row=0, column=2, padx=10, pady=10, sticky="w")
        self.tax_rate_entry = ctk.CTkEntry(self.params_frame)
        self.tax_rate_entry.insert(0, "0.06") # 默认6%
        self.tax_rate_entry.grid(row=0, column=3, padx=10, pady=10, sticky="ew")
        
        ctk.CTkLabel(self.params_frame, text="项目名称:").grid(row=0, column=4, padx=10, pady=10, sticky="w")
        self.item_name_entry = ctk.CTkEntry(self.params_frame)
        self.item_name_entry.insert(0, "餐饮服务") # 默认合并后的名称
        self.item_name_entry.grid(row=0, column=5, padx=10, pady=10, sticky="ew")

        # Action frame
        self.action_frame = ctk.CTkFrame(self)
        self.action_frame.grid(row=2, column=0, padx=20, pady=10, sticky="ew")
        self.action_frame.grid_columnconfigure(0, weight=1)

        self.run_button = ctk.CTkButton(self.action_frame, text="开始合并并生成", command=self.process_data)
        self.run_button.grid(row=0, column=0, padx=20, pady=20)

    def select_source(self):
        # 优化文件过滤器
        path = filedialog.askopenfilename(filetypes=[
            ("源数据文件 (xlsx/xls/csv)", "*.xlsx *.xls *.csv"),
            ("Excel 文件", "*.xlsx *.xls"),
            ("CSV 文件", "*.csv"),
            ("所有文件", "*.*")
        ])
        if path: self.source_path.set(path)

    def select_template(self):
        # 优化模板文件过滤器
        path = filedialog.askopenfilename(filetypes=[
            ("模板 Excel 文件 (xlsx)", "*.xlsx"),
            ("所有文件", "*.*")
        ])
        if path: self.template_path.set(path)

    def process_data(self):
        source_file = self.source_path.get()
        template_file = self.template_path.get()
        tax_code = self.tax_code_entry.get()
        tax_rate = self.tax_rate_entry.get()
        item_name = self.item_name_entry.get()

        if not source_file or not template_file:
            messagebox.showwarning("提示", "请先选择两个文件！")
            return

        try:
            # 1. 读取源数据
            if source_file.endswith('.csv'):
                df = pd.read_csv(source_file)
            else:
                df = pd.read_excel(source_file)

            # 数据预处理：确保字段为字符串，并处理空值
            df['金额'] = pd.to_numeric(df['金额'], errors='coerce').fillna(0)
            df['税号'] = df['税号'].fillna("").astype(str)
            df['公司主体'] = df['公司主体'].fillna("个人") 
            df['开票人'] = df['开票人'].fillna("").astype(str) # <-- 修复点1: 确保开票人（邮件）不为空
            
            # 转换时间以便排序
            df['创建时间'] = pd.to_datetime(df['创建时间'], errors='coerce')

            # 2. 核心逻辑：分组 (按开票人、公司、税号分组)
            grouped = df.groupby(['开票人', '公司主体', '税号'])
            
            # 3. 加载模板
            wb = openpyxl.load_workbook(template_file)
            sheet_basic = wb["1-发票基本信息"]
            sheet_detail = wb["2-发票明细信息"]

            # 起始行
            row_idx_basic = 4
            row_idx_detail = 4

            # 4. 遍历每个分组进行填入
            for i, ((person, company, tax_id), group_df) in enumerate(grouped):
                # --- 聚合计算 ---
                total_amount = group_df['金额'].sum() # 总金额
                
                # 如果总金额为0，跳过
                if total_amount == 0:
                    continue

                # 生成新的合并流水号
                first_order_id = str(group_df.iloc[0]['订单号'])
                if first_order_id.endswith(".0"): first_order_id = first_order_id[:-2]
                
                new_invoice_no = f"{first_order_id}_合"

                # 计算备注信息
                min_date = group_df['创建时间'].min().strftime('%m月%d日')
                max_date = group_df['创建时间'].max().strftime('%m月%d日')
                count = len(group_df)
                location = group_df.iloc[0]['消费地点'].split('-')[0]
                
                memo = f"{location} {min_date}-{max_date} {item_name}共{count}笔"

                # --- 填写 Sheet 1: 发票基本信息 ---
                sheet_basic.cell(row=row_idx_basic, column=1, value=new_invoice_no)     # 发票流水号 (A)
                sheet_basic.cell(row=row_idx_basic, column=2, value="增值税电子普通发票") # 发票类型 (B)
                sheet_basic.cell(row=row_idx_basic, column=4, value="是")               # 是否含税 (D)
                sheet_basic.cell(row=row_idx_basic, column=6, value=company)            # 购买方名称 (F)
                sheet_basic.cell(row=row_idx_basic, column=7, value=tax_id)             # 购买方纳税人识别号 (G)
                sheet_basic.cell(row=row_idx_basic, column=23, value=memo)              # 备注 (W)
                
                # 修复点2: 购买方电子邮箱 (AD列, 第30列)
                # 确保填入的person是一个有效的邮箱格式
                if '@' in person:
                    sheet_basic.cell(row=row_idx_basic, column=30, value=person)        # 购买方电子邮箱 (AD)

                row_idx_basic += 1

                # --- 填写 Sheet 2: 发票明细信息 ---
                sheet_detail.cell(row=row_idx_detail, column=1, value=new_invoice_no)   # 发票流水号 (A)
                sheet_detail.cell(row=row_idx_detail, column=2, value=item_name)        # 项目名称 (B)
                sheet_detail.cell(row=row_idx_detail, column=3, value=tax_code)         # 税收编码 (C)
                sheet_detail.cell(row=row_idx_detail, column=5, value="项")             # 单位 (E)
                sheet_detail.cell(row=row_idx_detail, column=6, value=1)                # 数量 (F)
                sheet_detail.cell(row=row_idx_detail, column=7, value=total_amount)     # 单价 (G)
                sheet_detail.cell(row=row_idx_detail, column=8, value=total_amount)     # 金额 (H)
                sheet_detail.cell(row=row_idx_detail, column=9, value=tax_rate)         # 税率 (I)

                row_idx_detail += 1
                
            # 5. 保存
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            output_folder = os.path.dirname(source_file)
            output_path = os.path.join(output_folder, f"已合并整理_开票文件_{timestamp}.xlsx")
            
            wb.save(output_path)
            
            CustomMessageBox(self, "成功", f"合并处理完成！\n共生成 {row_idx_basic - 4} 张发票数据。\n文件已保存至：\n{output_path}")

        except Exception as e:
            import traceback
            error_msg = traceback.format_exc()
            messagebox.showerror("错误", f"发生错误：{str(e)}\n\n{error_msg}")

if __name__ == "__main__":
    app = InvoiceToolApp()
    app.mainloop()